from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import multiprocessing as mp
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill, colors
import time

def thread(x,y,xl_data):
    driver = webdriver.Firefox(executable_path="C:\\Users\\Administrator\\Downloads\\geckodriver-v0.26.0-win64\\geckodriver.exe")
    driver.get("http://www.results.manabadi.co.in/2017/telangana/Inter-2nd/ts-intermediate-2nd-year-regular-exam-results-2017.htm")
    elem = driver.find_element_by_id("htno")
    select = Select(driver.find_element_by_id('Degree'))
    btn = driver.find_element_by_id("btnsubmit")
    hallTicket = driver.find_element_by_id("sid0")
    name = driver.find_element_by_id("sid2")
    data = []
    for i in range(x,y):
        elem.clear()
        elem.send_keys(i)
        select.select_by_visible_text('Inter 2Year')
        btn.click()
        time.sleep(0.7)
        if(name.text == ""):
            data.append(("Invalid","Nan"))
        else:
            #ht.value =
            #na.value = name.text
            data.append((hallTicket.text, name.text))
    xl_data+=[*data]
    driver.quit()

if __name__ == "__main__":
    xl_data = mp.Manager().list()

    t1 = mp.Process(target=thread,args = (1761216000,1761216500,xl_data))
    t2 = mp.Process(target=thread,args = (1761216501,1761217000,xl_data))
    t3 = mp.Process(target=thread,args = (1761217001,1761217500,xl_data))
    t4 = mp.Process(target=thread,args = (1761217501,1761218000,xl_data))

    t1.start()
    t2.start()
    t3.start()
    t4.start()

    t1.join()
    t2.join()
    t3.join()
    t4.join()
    wb = Workbook()
    sheet = wb.active
    j = 1
    for reg_no,name in xl_data:
        sheet.cell(row = j,column = 1).value = reg_no
        sheet.cell(row = j,column = 2).value = name
        j = j+1

    wb.save('scrappedData.xlsx')
    print("Done!")
